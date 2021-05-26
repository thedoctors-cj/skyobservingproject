import sys
import datetime
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment
from operator import itemgetter

##############################################################
# DOME - split this into 3 files?
# DOME - use calculate_acceptable_phases in the main GradeAnswers routine?
##############################################################
# HTML tag stripper class from StackExchange
##############################################################
from io import StringIO
from html.parser import HTMLParser

class MLStripper(HTMLParser):
    def __init__(self):
        super().__init__()
        self.reset()
        self.strict = False
        self.convert_charrefs= True
        self.text = StringIO()
    def handle_data(self, d):
        self.text.write(d)
    def get_data(self):
        return self.text.getvalue()

    
def strip_tags(html):
    s = MLStripper()
    s.feed(html)
    return s.get_data()


##############################################################
# Subroutines
##############################################################

# This routine cleans the blackboard info
def isolateAnswers(answers):
    # keep username, lastName and firstName
    just_answers = answers[0:3]
    answers = answers[3:]
    while (len(answers) > 0):
        answers = answers[2:]
        just_answers.append(answers.pop(0))
        answers = answers[3:]

    return just_answers


# This subroutine takes four parameters
# * the name of the student
# * a flag to determine whether the extra "teacher" output should be printed
# * the scores
# * the global variables
def summarizeResults(name, teacherOutput, scores, myGlobals):
    observation_number = myGlobals["observation_number"]
    if observation_number < 8:
        questions = myGlobals["questions"]
        questionsText = myGlobals["questionsText"]
        points = myGlobals["questionsPoints"]
    else:
        questions = myGlobals["summaryQuestions"]
        questionsText = myGlobals["summaryQuestionsText"]
        points = myGlobals["summaryPoints"]
        
    output = f"{scores[name]['lastName']}, {scores[name]['firstName']} ({name})\n\n"
    
    if teacherOutput:
        output = output +  "(Teacher Only) Blackboard Entries:\n"
        output = output +  f"\t{scores[name]['blackboardEntries']}\n\n"
    for ii in range(0,len(questions)):
        output = output +  f"{questionsText[ii]}:\n"
        output = output +  f"\tResponse: {scores[name][questions[ii]]['value']}\n"
        output = output +  f"\tPoints  : {scores[name][questions[ii]]['score']} of {points[ii]} points\n"
        if scores[name][questions[ii]]['comment']:
              output = output +  f"\tComments: {scores[name][questions[ii]]['comment']}\n"
        if teacherOutput and ("commentTeacher" in scores[name][questions[ii]]):
            output = output +  f"\t(Teacher Only) {scores[name][questions[ii]]['commentTeacher']}\n"
        output = output +  "\n"

    # Not sure why we check if the score exists and then just print the comment
    if "score" in scores[name]["moonModification"]:
        output = output +  f"{scores[name]['moonModification']['comment']}\n"
        output = output +  "\n"

    output = output +  f"\nTotal Points: {scores[name]['totalPoints']}\n"
    return output


def calculateTotalPoints(hash_ref, questions, verbose):
    hash_ref["totalPoints"] = 0
    for ii in range(0,len(questions)):
        hash_ref["totalPoints"] += hash_ref[questions[ii]]["score"]
    if "score" in hash_ref["moonModification"]:
        hash_ref["totalPoints"] += hash_ref["moonModification"]["score"]
    if verbose:
        print(f"Total Points: {hash_ref['totalPoints']}", file=sys.stderr)


def fillOutRemaining(hash_ref, reason, questions, verbose):
    for ii in range(0,len(questions)):
        if not "value" in hash_ref[questions[ii]]:
            hash_ref[questions[ii]]["value"] = ""
        if not "score" in hash_ref[questions[ii]]:
            hash_ref[questions[ii]]["score"] = 0
        if not "comment" in hash_ref[questions[ii]]:
            hash_ref[questions[ii]]["comment"] = reason
    hash_ref['saw_moon']['value'] = -1
    calculateTotalPoints(hash_ref, questions, verbose)


# The scores dictionary is keyed by the userID.  For each userID we
# fill in student info along with other info.
def createStudentMessages(scores, student_message_dir, myGlobals):
    sorted_userids = myGlobals["sorted_userids"]
    questions = myGlobals["questions"]
    questionsText = myGlobals["questionsText"]
    verbose = myGlobals["verbose"]
    for name in sorted_userids:
        if verbose:
            print(f"Working on {name}", file=sys.stderr)
        outfile = student_message_dir + "/" + name + "_" + student_message_dir + ".txt"
        try:
            with open(outfile,"w") as OUT:
                OUT.write(summarizeResults(name, 0, scores, myGlobals))
        except IOError as e:
            sys.exit(f"EXITING: {outfile} can't be opened")
        if verbose:
            print(f" Finished Working on {name}", file=sys.stderr)


def createTeacherSummary(scores, flagged, student_message_dir, myGlobals):
    printDaylightSavingsCheck = myGlobals["printDaylightSavingsCheck"]
    daylight_savings_check = myGlobals["daylight_savings_check"]
    sorted_userids = myGlobals["sorted_userids"]
    questions = myGlobals["questions"]
    questionsText = myGlobals["questionsText"]
    verbose = myGlobals["verbose"]
    if verbose:
        print(f" Working on teacher summary", file=sys.stderr)
    outfile = student_message_dir + ".summary.withComments.txt"
    try:
        with open(outfile,"w") as OUT:
            OUT.write(f"Flagged the following issues:\n")
            for name in flagged.keys():
                OUT.write(f"{name} - {flagged[name]}\n")
            OUT.write(f"\n\n")
    
            if printDaylightSavingsCheck:
                for name in daylight_savings_check.keys():
                    OUT.write(f"{name} - Verify Daylight Savings\n")
                OUT.write(f"\n\n")

            for name in sorted_userids:
                OUT.write("----------------------------------------------------\n")
                OUT.write(summarizeResults(name, 1, scores, myGlobals))

    except IOError as e:
        sys.exit(f"EXITING: {outfile} can't be opened")
    if verbose:
        print(f"  Finished Working on teacher summary", file=sys.stderr)


def createTeacherSummaryNoTeacherComments(scores, flagged, student_message_dir, myGlobals):
    sorted_userids = myGlobals["sorted_userids"]
    questions = myGlobals["questions"]
    questionsText = myGlobals["questionsText"]
    verbose = myGlobals["verbose"]
    if verbose:
        print(f" Working on teacher summary", file=sys.stderr)
    outfile = student_message_dir + ".summary.noComments.txt"
    try:
        with open(outfile,"w") as OUT:
            OUT.write(f"Flagged the following issues:\n")
            for name in flagged.keys():
                OUT.write(f"{name} - {flagged[name]}\n")
            OUT.write(f"\n\n")

            for name in sorted_userids:
                OUT.write("----------------------------------------------------\n")
                OUT.write(summarizeResults(name, 0, scores, myGlobals))
    except IOError as e:
        sys.exit(f"EXITING: {outfile} can't be opened")
    if verbose:
        print(f"  Finished Working on teacher summary", file=sys.stderr)


def printGrades(scores, student_message_dir, myGlobals):
    sorted_userids = myGlobals["sorted_userids"]
    outfile = student_message_dir + ".grades.txt"
    try:
        with open(outfile,"w") as OUT:
            for name in sorted_userids:
                OUT.write(f"{scores[name]['lastName']}, {scores[name]['firstName']} ({name})\t{float(scores[name]['totalPoints']):g}\n")
    except IOError as e:
        sys.exit(f"EXITING: {outfile} can't be opened")


def load_sunset_times(sunset_times, myGlobals):
    sunrise_sunset_dir = myGlobals["sunrise_sunset_dir"]
    year = myGlobals["year"]
    sunset_file = f"{sunrise_sunset_dir}/{year}.sunset"
    try:
        with open(sunset_file,"r") as SUNSETS:
            for line in SUNSETS:
                line = line.rstrip()
                date, time = line.split("\t")
                sunset_times[datetime.datetime.strptime(date, '%m/%d/%Y')] = time
    except IOError as e:
        sys.exit(f"EXITING: {sunset_file} can't be opened")


def dequote(s):
    """
    If a string has single or double quotes around it, remove them.
    Make sure the pair of quotes match.
    If a matching pair of quotes is not found, return the string unchanged.
    """
    if len(s) == 0:
        return s
    if (s[0] == s[-1]) and s.startswith(("'", '"')):
        return s[1:-1]
    return s


def printSpreadsheet(scores, student_message_dir, myGlobals):
    sorted_userids = myGlobals["sorted_userids"]
    questions = myGlobals["questions"]
    questionsText = myGlobals["questionsText"]
    outfile = student_message_dir + ".grades.xlsx"
    workbook = Workbook()
    ws = workbook.active
    # Need to add this to each cell I create
    # ws['A1'].alignment = Alignment(wrapText=True)
    row = 1
    print(f"Doing Spreadsheet", file=sys.stderr)
    # DOME - Need to figure out how to sorty the userIDs by lastName
    # for name in  (sort { $scores{$a}{lastName} cmp $scores{$b}{lastName} } keys %scores)
    for name in sorted_userids:
        cell = "A" + str(row)
        ws[cell] = f"{scores[name]['lastName']}, {scores[name]['firstName']}"
        ws[cell].alignment = Alignment(wrapText=True)
        cell = "B" + str(row)
        ws[cell] = scores[name]['totalPoints']
        ws[cell].alignment = Alignment(wrapText=True)

        comments = summarizeResults(name, 0, scores, myGlobals);
        # This seems to improve the spreadsheet readability
        comments = comments.replace("\t", "    ")
        cell = "C" + str(row);
        ws[cell] = comments
        ws[cell].alignment = Alignment(wrapText=True)
        row += 1
    workbook.save(outfile)

    
def createTestData(answers):
    print(f"{answers[0]}\t")
    for ii in range(1,len(answers), 3):
        print(f"{answers[ii]}\t")
    print("\n")

def fixDegreesNorthSouth(deg):
    if deg < 0:
        deg *= -1
        deg = f"{deg} South"
    else:
        deg = f"{deg} North"
    return deg

# This subroutine takes in the month, day and year and returns 1 if it
# is during daylight savings time, and zero otherwise
def daylight_savings(month, day, year, myGlobals):
    daylight_savings_start = myGlobals["daylight_savings_start"]
    daylight_savings_end = myGlobals["daylight_savings_end"]
    if month < 3:
        return 0
    if month > 11:
        return 0
    if ( (month > 3) and (month < 11) ):
        return 1 
    if month == 3:
        if day < daylight_savings_start[year]:
            return 0
        else:
            return 1
    if month == 11:
        if day >= daylight_savings_end[year]:
            return 0
        else:
            return 1
    sys.exit(f"EXITING: Could not figure out daylight savings!")

# Just load the name and date from the previous observations
def load_previous_observation_dates(previous_observation_dates_file, previous_observation_dates):
    try:
        with open(previous_observation_dates_file,"r") as IN:
            for line in IN:
                line = line.rstrip()
                name_date = line.split("\t")
                previous_observation_dates[name_date[0]] = datetime.datetime.strptime(name_date[1], '%Y-%m-%d %H:%M:%S')
    except IOError as e:
        sys.exit(f"EXITING: {previous_observation_dates_file} can't be opened")

def is_float(my_num):
    try:
        float(my_num)
        return True
    except:
        return False
        
# Load all entries from the previous observations
def load_all_previous_observation_dates(myGlobals):
    verbose = myGlobals["verbose"]
    if verbose:
        print(f"Loading Previous Observation Dates", file=sys.stderr)
    for ii in range(1,8):
        previous_observation_dates_file = "../" + myGlobals["previous_observation_dates_basename"] + str(ii)
        if verbose:
            print(f"  Loading File: {previous_observation_dates_file}", file=sys.stderr)
        try:
            with open(previous_observation_dates_file,"r") as IN:
                for line in IN:
                    line = line.rstrip()
                    name_date = line.split("\t")
                    myGlobals["summary_all"][name_date[0]]["LastObservationDate"] = name_date[1]
                    myGlobals["summary_all"][name_date[0]][name_date[1]]["num"] = ii
                    # Before we save values, we verify that the file has suitable values stored
                    if is_float(name_date[2]):
                        myGlobals["summary_all"][name_date[0]][name_date[1]]["DegreesNorth"] = float(name_date[2])
                    else:
                        myGlobals["summary_all"][name_date[0]][name_date[1]]["DegreesNorth"] = -1
                    if name_date[3].isdecimal():
                        myGlobals["summary_all"][name_date[0]][name_date[1]]["SawMoon"] = int(name_date[3])
                    else:
                        myGlobals["summary_all"][name_date[0]][name_date[1]]["SawMoon"] = -1
                    myGlobals["summary_all"][name_date[0]][name_date[1]]["MoonShouldHaveBeenVisible"] = name_date[4]
                    myGlobals["summary_all"][name_date[0]][name_date[1]]["ObservationWasZeroed"] = name_date[5]
        except IOError as e:
            sys.exit(f"EXITING: {previous_observation_dates_file} can't be opened")
    if verbose:
        print(f"Done Loading Previous Observation Dates", file=sys.stderr)




        
# Make sure to save off these observation dates
def save_current_observation_dates(current_observation_dates_file, scores, zeroed_score, moon_visible):
    dates_outfile = current_observation_dates_file
    try:
        with open(dates_outfile,"w") as OBS:
            for name in scores.keys():
                if scores[name]["date"]["comment"] == "Date out of Range":
                    continue
                if len(scores[name]["date"]["value"]) == 0:
                    continue
                try:
                    this_observation_date = datetime.datetime.strptime(scores[name]["date"]["value"], '%m/%d/%Y')
                except:
                    continue
                # need to add phase
                OBS.write(f"{name}\t")
                OBS.write(f"{this_observation_date}\t")
                OBS.write(f"{scores[name]['degrees_north']['value']}\t")
                OBS.write(f"{scores[name]['saw_moon']['value']}\t")
                OBS.write(f"{moon_visible[name]}\t")
                this_was_zeroed = 0
                if name in zeroed_score:
                    this_was_zeroed = 1
                OBS.write(f"{this_was_zeroed}\t")
                OBS.write(f"\n")

    except IOError as e:
        sys.exit(f"EXITING: {dates_outfile} can't be opened")


def getSunAz(date, h, m, alt_azm_dir, verbose):
    year = date.strftime("%Y")
    month = date.strftime("%m")
    day = date.strftime("%d")
    sun_file = f"{alt_azm_dir}/{year}/sun_{year}{month}{day}.txt"
    time = str(h+12) + ":" + str(m).zfill(2)
    sun_azimuth = -1
    try:
        with open(sun_file,"r") as FILE:
            if verbose:
                print(f" File {sun_file} opened, trying to match {time}", file=sys.stderr)
            for line in FILE:
                if not time in line:
                    continue
                sun_azimuth = re.sub(rf"^{time}.+\s+(\d+\.\d)$", r"\1",line).rstrip()
                if verbose:
                    print(f" Matched the line {line}", file=sys.stderr)
                break

    except IOError as e:
        sys.exit(f"EXITING: {sun_file} can't be opened")

    if verbose:
        print(f" Sun azimuth = {sun_azimuth}", file=sys.stderr)
    return float(sun_azimuth)


def getMoonData(date, h, m, myGlobals):
    alt_azm_dir = myGlobals["alt_azm_dir"]
    verbose = myGlobals["verbose"]
    year = date.strftime("%Y")
    month = date.strftime("%m")
    day = date.strftime("%d")
    moon_file = f"{alt_azm_dir}/{year}/moon_{year}{month}{day}.txt";
    time = str(h+12) + ":" + str(m).zfill(2)
    moon_altitude = -1
    moon_azimuth = -1
    moon_max_altitude = -1
    fraction_illuminated = -1
    fraction_illuminated_start = -1
    fraction_illuminated_end = -1
    try:
        with open(moon_file,"r") as FILE:
            if verbose:
                print(f" File {moon_file} opened, trying to match {time}", file=sys.stderr)
            for line in FILE:
                my_match = re.match(r'^(\d\d:\d\d)\s+(\S+)\s+(\S+)\s+(\S+)$',line)
                if my_match:
                    time_tmp = my_match.group(1)
                    alt_tmp = float(my_match.group(2))
                    az_tmp = float(my_match.group(3))
                    frac_tmp = float(my_match.group(4))
                    if alt_tmp > moon_max_altitude:
                        moon_max_altitude = alt_tmp
                    if fraction_illuminated_start == -1:
                        fraction_illuminated_start = frac_tmp 
                    fraction_illuminated_end = frac_tmp
                    if time == time_tmp:
                        moon_altitude = alt_tmp
                        moon_azimuth = az_tmp
                        fraction_illuminated = frac_tmp
                        if verbose:
                            print(f" Matched the line {line}", file=sys.stderr)

    except IOError as e:
        sys.exit(f"EXITING: {moon_file} can't be opened")

    moon_change = ""
    if fraction_illuminated_start < fraction_illuminated_end:
        moon_change = 1
    elif fraction_illuminated_start > fraction_illuminated_end:
        moon_change = -1
    else:
        moon_change = 0
    fraction_illuminated_ave = (fraction_illuminated_start + fraction_illuminated_end)/2
    if verbose:
        print(f" Moon Altitude: {moon_altitude}", file=sys.stderr)
        print(f" Moon Max Altitude: {moon_max_altitude}", file=sys.stderr)
        print(f" Moon Azimuth: {moon_azimuth}", file=sys.stderr)
        print(f" Fraction Illuminated: {fraction_illuminated}", file=sys.stderr)
        print(f" Fraction Illuminated at start: {fraction_illuminated_start}", file=sys.stderr)
        print(f" Fraction Illuminated at end  : {fraction_illuminated_end}", file=sys.stderr)
        print(f" Moon waning/nothing/waxing -1/0/1: {moon_change}", file=sys.stderr)
    return moon_altitude, moon_max_altitude, fraction_illuminated, moon_change, moon_azimuth, fraction_illuminated_ave

# WARNING - this isn't quite the same as the one from the
# gradeObservations routine since I removed the variable
# $expected_moon_visible.  If you add an outer check against that
# variable this could be combined into that one later.
def calculate_acceptable_phases(date, myGlobals):
    moon_phase_slop = myGlobals['moon_phase_slop']
    moon_altitude, moon_max_altitude, fraction_illuminated, moon_change, moon_azimuth, fraction_illuminated_ave = getMoonData(date, 11, 59, myGlobals)
    acceptablePhase = {}
    
    if fraction_illuminated_ave<=0: 
        acceptablePhase[8] = 1
    else:
        if((moon_change > 0) and (fraction_illuminated_ave < (.5+moon_phase_slop))):
            acceptablePhase[1] = 1 
        if((moon_change > 0) and (fraction_illuminated_ave < (.5+moon_phase_slop)) and (fraction_illuminated_ave > (.5-moon_phase_slop))):
            acceptablePhase[2] = 1 
        if((moon_change > 0) and (fraction_illuminated_ave > (.5-moon_phase_slop))):
            acceptablePhase[3] = 1 
        if((fraction_illuminated_ave > (1-moon_phase_slop))):
            acceptablePhase[4] = 1 
        if((moon_change < 0) and (fraction_illuminated_ave > (.5-moon_phase_slop))):
            acceptablePhase[5] = 1 
        if((moon_change < 0) and (fraction_illuminated_ave < (.5+moon_phase_slop)) and (fraction_illuminated_ave > (.5-moon_phase_slop))):
            acceptablePhase[6] = 1 
        if((moon_change < 0) and (fraction_illuminated_ave < (.5+moon_phase_slop))):
            acceptablePhase[7] = 1 
        if((fraction_illuminated_ave < (moon_phase_slop))):
            acceptablePhase[8] = 1 

    comment = ""
    if(moon_change == 1):
        comment = comment + "Waxing "
    if(moon_change == -1):
        comment = comment + "Waning " 
    comment = comment + f"{fraction_illuminated_ave} Illuminated";
    print(comment, file=sys.stderr)
    acceptablePhase["comment"] = comment
    return acceptablePhase




# This subroutine takes a list of answers and grades them
#
# - For each possible question, we fill in value/score/comment (even if the comment is the null string)
# - We optionally fill in a "commentTeacher" parameter which is only output to the teacher
# - we also fill in a "flagged" hash table that is output to the teacher for followon analysis
def gradeObservations(answers, scores, flagged, previous_observation_dates, sunset_times, moon_visible, zeroed_score, myGlobals):
    year = myGlobals["year"]
    start_date = myGlobals["start_date"]
    end_date = myGlobals["end_date"]
    observation_number = myGlobals["observation_number"]
    questions = myGlobals["questions"]
    daylight_savings_start = myGlobals["daylight_savings_start"]
    daylight_savings_check = myGlobals["daylight_savings_check"]
    moon_altitude_min = myGlobals["moon_altitude_min"]
    alt_azm_dir = myGlobals["alt_azm_dir"]
    reverseNamePhase = myGlobals["reverseNamePhase"]
    skyLocation = myGlobals["skyLocation"]
    verbose = myGlobals["verbose"]
    sun_degree_slop = myGlobals["sun_degree_slop"]
    moon_phase_slop = myGlobals["moon_phase_slop"]
    moon_location_slop = myGlobals["moon_location_slop"]
    moon_location_cutoff = myGlobals["moon_location_cutoff"]
    reverseSkyLocation = myGlobals["reverseSkyLocation"]
     # keep their name with their score array
    name = answers.pop(0)
    scores[name]["lastName"] = answers.pop(0)
    scores[name]["firstName"] = answers.pop(0)
    # Note: need to use copy() in order to maintain the original list
    scores[name]["blackboardEntries"] = answers.copy()

    if verbose:
        print(f"{scores[name]['lastName']}, {scores[name]['firstName']} ({name})", file=sys.stderr)
        print(f" Blackboard Entries: {scores[name]['blackboardEntries']}", file=sys.stderr)
    # If this is the first observation, we treat the first three answers separately
    if observation_number == 1:
        scores[name]["howDetermineDirection"]["value"] = answers.pop(0)
        scores[name]["explainOther"]["value"] = answers.pop(0)
        scores[name]["whereObservation"]["value"] = answers.pop(0)
        scores[name]["howDetermineDirection"]["value"] = scores[name]["howDetermineDirection"]["value"].replace("<Unanswered>","")
        scores[name]["explainOther"]["value"] = scores[name]["explainOther"]["value"].replace("<Unanswered>","")
        scores[name]["whereObservation"]["value"] = scores[name]["whereObservation"]["value"].replace("<Unanswered>","")

        # get a point for how to determine direction unless they pick
        # other and then don't put in an explanation
        if "Other" in scores[name]["howDetermineDirection"]:
            if len(scores[name]["explainOther"]["value"]) == 0:
                scores[name]["howDetermineDirection"]["score"] = 0
                scores[name]["howDetermineDirection"]["comment"] = "Nothing given to explain how the direction was determined"
            else:
                scores[name]["howDetermineDirection"]["score"] = 1
                scores[name]["howDetermineDirection"]["comment"] = f"{scores[name]['explainOther']['value']} used to determine direction"
        else:
            scores[name]["howDetermineDirection"]["score"] = 1;
            scores[name]["howDetermineDirection"]["comment"] = ""

        if verbose:
            print(f" Obs Direction Determination: {scores[name]['howDetermineDirection']['value']} {scores[name]['howDetermineDirection']['score']} point {scores[name]['howDetermineDirection']['comment']}", file=sys.stderr)

        if len(scores[name]['whereObservation']['value']) == 0:
            scores[name]['whereObservation']['score'] = 0;
            scores[name]['whereObservation']['comment'] = "No observation location given"
        else:
            scores[name]['whereObservation']['score'] = 1
            scores[name]['whereObservation']['comment'] = ""

        if verbose:
            print(f" Obs Location: {scores[name]['whereObservation']['value']} {scores[name]['whereObservation']['score']} point {scores[name]['whereObservation']['comment']}", file=sys.stderr)

    # Next entry is a date - clean and store, then verify
    this_observation_month = answers.pop(0)
    this_observation_day = answers.pop(0)
    scores[name]["date"]["value"] = f"{this_observation_month}/{this_observation_day}/{year}"
    scores[name]["date"]["value"] = scores[name]["date"]["value"].replace("\"","")
    scores[name]["date"]["value"] = scores[name]["date"]["value"].replace(" ","")
    scores[name]["date"]["value"] = scores[name]["date"]["value"].replace(".","")
    try:
        this_observation_date = datetime.datetime.strptime(scores[name]["date"]["value"], '%m/%d/%Y')
        this_observation_month = this_observation_date.month
        this_observation_day = this_observation_date.day
    except:
        # if the observation date is not properly converted, flag and bail
        flagged[name] = "Date could not be converted";
        fillOutRemaining(scores[name], "No points - could not recognize date", questions, verbose)
        return(-1)

    if verbose:
        print(f" Obs date: {this_observation_month}/{this_observation_day}/{year} converted to {this_observation_date}", file=sys.stderr)

    # Check that the observation is within the required date range if
    # not, no points and bail - they get nothing for this period.
    problem_with_observation_date = False
    if (this_observation_date-start_date).total_seconds() < 0:
        scores[name]["date"]["score"] = 0;
        scores[name]["date"]["comment"] = "Observation date was before the start of the observation period";
        problem_with_observation_date = True
    if (end_date-this_observation_date).total_seconds() < 0:
        scores[name]["date"]["score"] = 0;
        scores[name]["date"]["comment"] = "Observation date was after the end of the observation period";
        problem_with_observation_date = True
    if problem_with_observation_date:
        flagged[name] = scores[name]["date"]["comment"]
        fillOutRemaining(scores[name], f"No points - {scores[name]['date']['comment']}", questions, verbose)
        if verbose:
            print(f" Date: {scores[name]['date']['score']} point {scores[name]['date']['comment']}", file=sys.stderr)
        return(-1)

    # We get here if the observation was correctly made in the observation period
    # If there is no record of a past observation, they get the point but it is commented
    if not name in previous_observation_dates:
        scores[name]["date"]["score"] = 1;
        if observation_number == 1:
            scores[name]["date"]["comment"] = "First observation - no previous observation"
        else:
            scores[name]['date']['comment'] = "No record of previous observation"
    # we get here if there was a previous observation.  Now we make
    # sure it has been more than a week since the last one
    elif (this_observation_date-previous_observation_dates[name]-datetime.timedelta(days=7)).total_seconds() < 0:
        scores[name]["date"]["score"] = 0
        scores[name]["date"]["comment"] = "Observation not at least a week since last one";
    else:
        scores[name]["date"]["score"] = 1
        scores[name]["date"]["comment"] = ""
    if verbose:
        print(f" Date: {scores[name]['date']['score']} point {scores[name]['date']['comment']}", file=sys.stderr)

    # 1 observation time = hr/min:  check within correct range-- no less
    # than 10 min no more than 45 min before USNO sunset time  (no= -1pt)
    # IF AFTER SUNSET TIME, OR MORE THAN 1 HOUR BEFORE SUNSET, NO CREDIT FOR OBSERVATION
    this_observation_hour = answers.pop(0)
    this_observation_minute = answers.pop(0)
    this_observation_time_original = f"Hour: {this_observation_hour}, Minute: {this_observation_minute}";
    
    # Here we do some cleaning on the time to make sure it is
    #  if the observation minute has only one digit, we fix that here
    this_observation_hour   = re.sub(r'.*?(\d+).*',r'\1', this_observation_hour) # grab the first set of digits
    this_observation_minute = re.sub(r'.*?(\d+).*',r'\1', this_observation_minute) # grab the first set of digits
    can_parse_time = True
    try:
        tmp = int(this_observation_hour)
    except:
        can_parse_time = False
    try:
        tmp = int(this_observation_minute)
    except:
        can_parse_time = False
    if not can_parse_time:
        scores[name]["time"]["score"] = 0
        scores[name]["time"]["comment"] = f"Could not determine time from {this_observation_time_original}";
        if verbose:
            print(f" Time: {scores[name]['time']['score']} point {scores[name]['time']['comment']}", file=sys.stderr)
        flagged[name] = scores[name]["time"]["comment"]
        fillOutRemaining(scores[name], f"No points - {scores[name]['time']['comment']}", questions, verbose)
        return(-1)
    else:
        this_observation_hour   = this_observation_hour.zfill(2)
        this_observation_minute = this_observation_minute.zfill(2)

    # Save off the time observed by the student - this may be EST or
    # EDT depending on the calendar but we adjust below
    scores[name]["time"]["value"] = f"{this_observation_hour}:{this_observation_minute}";

    # We get here if we have parsable dates and times.  Now we check
    # if we are in daylight savings time.  If so, we subtract one hour
    # from the observation time to get back to Eastern Standard time.
    if daylight_savings(int(this_observation_month), int(this_observation_day), year, myGlobals):
        daylight_savings_check[name] = 1
        scores[name]["daylightSavings"]["value"] = 1
        this_observation_hour = int(this_observation_hour)-1
    else:
        scores[name]["daylightSavings"]["value"] = 0

    official_sunset = sunset_times[this_observation_date]
    # The following is just to output nicely for the user - DST_adjusted means it is modified based on DST (not that it is a DST value)
    official_sunset_DST_adjusted = str(int(official_sunset[0:2])+scores[name]["daylightSavings"]["value"]) + ":" + official_sunset[2:4]
    official_minutes_since_midnight = int(official_sunset[0:2])*60 + int(official_sunset[2:4])
    observed_minutes_since_midnight = 720 + int(this_observation_hour)*60 + int(this_observation_minute)
    if verbose:
        print(f" Offcial Sunset Time (mins ): {official_minutes_since_midnight}, Student Obs Time (mins ): {observed_minutes_since_midnight}", file=sys.stderr)
    if official_minutes_since_midnight-observed_minutes_since_midnight > 45:
        scores[name]["time"]["score"] = 0
        scores[name]["time"]["comment"] = f"Observation must be within 45 minutes of sunset. Sunset was {official_sunset_DST_adjusted}."
        if official_minutes_since_midnight-observed_minutes_since_midnight > 60:
            scores[name]["date"]["score"] = 0
            scores[name]["time"]["comment"] = scores[name]["time"]["comment"] + "No points for entire observation.  Zeroed out date points.";
            zeroed_score[name] = 1
            flagged[name] = scores[name]["time"]["comment"]
            if verbose:
                print(f" Time: {scores[name]['time']['score']} point {scores[name]['time']['comment']}", file=sys.stderr)
            fillOutRemaining(scores[name], f"No points - {scores[name]['time']['comment']}", questions, verbose)
            return(-1)
    elif official_minutes_since_midnight-observed_minutes_since_midnight < 10:
        scores[name]["time"]["score"] = 0
        scores[name]["time"]["comment"] = f"Observation should be at least 10 minutes before sunset.  Sunset was {official_sunset_DST_adjusted}.  ";
        if official_minutes_since_midnight-observed_minutes_since_midnight < 0:
            scores[name]["date"]["score"] = 0
            scores[name]["time"]["comment"] = scores[name]["time"]["comment"] + "No points for entire observation - observation made after sunset.  Zeroed out date points.";
            zeroed_score[name] = 1
            flagged[name] = scores[name]["time"]["comment"]
            if verbose:
                print(f" Time: {scores[name]['time']['score']} point {scores[name]['time']['comment']}", file=sys.stderr)
            fillOutRemaining(scores[name], f"No points - {scores[name]['time']['comment']}", questions, verbose)
            return(-1)
    else:
        scores[name]["time"]["score"] = 1
        scores[name]["time"]["comment"] = "Accepted " + str(official_minutes_since_midnight-observed_minutes_since_midnight) + f" minutes from sunset ({official_sunset_DST_adjusted})"
        if verbose:
            print(f" Time: {scores[name]['time']['score']} point {scores[name]['time']['comment']}", file=sys.stderr)

    # For the next two questions we grab the azimuth of the sun at the
    # observation time from the alt/az table downloaded
    sun_azimuth = getSunAz(this_observation_date, int(this_observation_hour), int(this_observation_minute), alt_azm_dir, verbose)
    official_degrees_north = sun_azimuth - 270

    # 1 T/F The sun set N of due W:   T if between Mar 20 2011 - Sep 23, 2011
    scores[name]["sun_set_north_tf"]["value"] = answers.pop(0)
    scores[name]["sun_set_north_tf"]["value"] = scores[name]["sun_set_north_tf"]["value"].lower()
    official_sun_set_north_tf = "true"
    if official_degrees_north < 0:
        official_sun_set_north_tf = "false"
    if verbose:
        print(f" Sun Set North of West T/F: expecting {official_sun_set_north_tf}", file=sys.stderr)
    if scores[name]["sun_set_north_tf"]["value"] == official_sun_set_north_tf:
        scores[name]["sun_set_north_tf"]["score"] = 1
        scores[name]["sun_set_north_tf"]["comment"] = ""
    else:
        scores[name]["sun_set_north_tf"]["score"] = 0
        scores[name]["sun_set_north_tf"]["comment"] = "Incorrect T/F"
    if verbose:
        print(f" Sun Set North of West T/F: {scores[name]['sun_set_north_tf']['score']} point {scores[name]['sun_set_north_tf']['comment']}", file=sys.stderr)

    # Within 10 degrees of true position by 3rd observation
    observed_degrees_north = answers.pop(0)
    scores[name]["degrees_north"]["value"] = observed_degrees_north
    my_match = re.match(r'.*?(\d+).*', observed_degrees_north)
    if not my_match:
        scores[name]["degrees_north"]["score"] = 0;
        scores[name]["degrees_north"]["comment"] = f"Can not find any degrees in this answer: {scores[name]['degrees_north']['value']}"
    else:
        observed_degrees_north = float(my_match.group(1))
        if observed_degrees_north > 90:
            scores[name]["degrees_north"]["score"] = 0
            scores[name]["degrees_north"]["comment"] = f"Degrees must be a number less than 90: {scores[name]['degrees_north']['value']}"
        else:
            scores[name]["degrees_north"]["value"] = observed_degrees_north
            if scores[name]["sun_set_north_tf"]["value"] == "false":
                observed_degrees_north *= -1 
            difference_degrees_north = official_degrees_north-observed_degrees_north

            # If they are within the correct number of degrees of the official
            # sun location, they get points for both even if they had the T/F
            # part incorrect
            if abs(difference_degrees_north) < sun_degree_slop[observation_number]:
                if not scores[name]["sun_set_north_tf"]["score"]:
                    scores[name]["sun_set_north_tf"]["score"] = 1
                    scores[name]["sun_set_north_tf"]["comment"] = "Incorrect T/F, but degrees were close enough to get points anyway!"
                    if verbose:
                        print(f" Sun Set North of West T/F: Changed score due to correct degree calculation", file=sys.stderr)
            

                scores[name]["degrees_north"]["score"] = 1
                scores[name]["degrees_north"]["comment"] = "Accepted " + fixDegreesNorthSouth(observed_degrees_north) + " for official position " + fixDegreesNorthSouth(official_degrees_north)

            else:
                scores[name]["degrees_north"]["score"] = 0
                scores[name]["degrees_north"]["comment"] = "Sun degrees off by " + str(abs(difference_degrees_north)) + " degrees"

    if verbose:
        print(f" Sun Set Degrees North of West: {scores[name]['degrees_north']['score']} point {scores[name]['degrees_north']['comment']}", file=sys.stderr)


    moon_altitude, moon_max_altitude, fraction_illuminated, moon_change, moon_azimuth, fraction_illuminated_ave = getMoonData(this_observation_date, int(this_observation_hour), int(this_observation_minute), myGlobals)

    # Moon was visible (look up based on moon rise time*) (one point)
    # - if no and should be no point
    # - IF YES WHEN MOON NOT UP, FLAG OBSERVATION
    # - acceptable to see Moon 10 min. after moonrise time - if student says
    #   "yes" for time between moonrise and moonrise +10min. then no point + flag
    #   ...
    #   For explanation of moon_altitude_min, see global variable declaration

    # TODO: Add the moon_visible_slop here to handle
    scores[name]["moon_visible_tf"]["value"] = answers.pop(0)
    scores[name]["moon_visible_tf"]["value"] = scores[name]["moon_visible_tf"]["value"].lower()
    expected_moon_visible = "true"
    if moon_altitude < moon_altitude_min:
        expected_moon_visible = "false"
    if fraction_illuminated <= 0:
        expected_moon_visible = "false" 

    if not (scores[name]['moon_visible_tf']['value'] == "true" or scores[name]['moon_visible_tf']['value'] == "false"):
        scores[name]['moon_visible_tf']['score'] = 0
        scores[name]['moon_visible_tf']['comment'] = "Neither true nor false answer given"
    else:
        if scores[name]['moon_visible_tf']['value'] == expected_moon_visible:
            scores[name]['moon_visible_tf']['score'] = 1
            scores[name]['moon_visible_tf']['comment'] = f"Correctly said {scores[name]['moon_visible_tf']['value']}"
        elif scores[name]['moon_visible_tf']['value'] == "false":
            # we get here when expected_moon_visible = "true"
            scores[name]['moon_visible_tf']['score'] = 0
            scores[name]['moon_visible_tf']['comment'] = "Should have seen the moon but didn't"
        else:
            # we get here when expected_moon_visible = "false"
            # check if the moon was really low - if so, we change the
            # expected_moon_visible value and give points
            if (moon_altitude > 0) and (moon_altitude < moon_altitude_min) and (fraction_illuminated > 0):
                expected_moon_visible = "true"
                scores[name]['moon_visible_tf']['score'] = 1
                scores[name]['moon_visible_tf']['comment'] = "Correctly said {scores[name]['moon_visible_tf']['value']}"
            else:
                scores[name]['moon_visible_tf']['score'] = 0
                scores[name]['moon_visible_tf']['comment'] = "Moon would not have been visible on this date."
                flagged[name] = scores[name]['moon_visible_tf']['comment']
                # RELAX - commented out the followng line for RELAXED grading
                # fillOutRemaining(scores[name], f"No points - Claimed to have seen the moon when it wasn't out", questions, verbose)
                # return(-1)
    if verbose:
        print(f" Moon Visible T/F: {scores[name]['moon_visible_tf']['score']} point {scores[name]['moon_visible_tf']['comment']}", file=sys.stderr)
    if expected_moon_visible == "true":
        moon_visible[name] = 1
    else:
        moon_visible[name] = 0

    # Match phase (look up based on % illuminated): (two points)
    # * if no and should be then no 2 points
    #   new         0% - 8
    #   crescent 0-49% - 1 or 7
    #   quarter    50% - 2 or 6
    #   gibbous 50-99% - 3 or 5
    #   full      100% - 4
    # * allow +/- 10% on matching phase

    # Added a question that allowed the user to upload an image which
    # we don't grade.  This happened after we used this script awhile,
    # so I'll test old years to see if it was the case.  Otherwise,
    # dumping the answer between "Moon was visible" and "Moon Phase"
    # because it is now a question where people just upload their
    # images
    if year > 2017:
        scores[name]['moon_phase']['value'] = answers.pop(0)
    scores[name]['moon_phase']['value'] = answers.pop(0)

    # First we verify that we got an answer between 1 and 8
    if not re.match(r'^[1-8]$',scores[name]['moon_phase']['value']):
        scores[name]['moon_phase']['score'] = 0
        scores[name]['moon_phase']['comment'] = "Answer not a number from 1 to 8"
    else:
        scores[name]['moon_phase']['value'] = int(scores[name]['moon_phase']['value'])
        scores[name]['moon_phase']['valueText'] = reverseNamePhase[scores[name]['moon_phase']['value']]
        # We add keys to this for only those acceptable phases
        acceptablePhase = {}

        # fraction_illuminated will be -1 if it wasn't in the table - not visible!
        if (fraction_illuminated <= 0) or (expected_moon_visible == "false"):
            acceptablePhase[8] = 1
        else:
            if ((moon_change > 0) and (fraction_illuminated<(.5+moon_phase_slop))):
                acceptablePhase[1] = 1 
            if ((moon_change > 0) and (fraction_illuminated<(.5+moon_phase_slop)) and (fraction_illuminated>(.5-moon_phase_slop))):
                acceptablePhase[2] = 1 
            if ((moon_change > 0) and (fraction_illuminated>(.5-moon_phase_slop))):
                acceptablePhase[3] = 1 
            if ((fraction_illuminated>(1-moon_phase_slop))):
                acceptablePhase[4] = 1 
            if ((moon_change<0) and (fraction_illuminated>(.5-moon_phase_slop))):
                acceptablePhase[5] = 1 
            if ((moon_change<0) and (fraction_illuminated<(.5+moon_phase_slop)) and (fraction_illuminated>(.5-moon_phase_slop))):
                acceptablePhase[6] = 1 
            if ((moon_change<0) and (fraction_illuminated<(.5+moon_phase_slop))):
                acceptablePhase[7] = 1 

            scores[name]['moon_phase_acceptable']['value'] = ""
            for phase in acceptablePhase.keys():
                scores[name]['moon_phase_acceptable']['value'] = scores[name]['moon_phase_acceptable']['value'] + f":{phase}"

        # flag observations where false for moon visible but quote a
        # moon phase from picture and a moon phase name
        if( (scores[name]['moon_visible_tf']['value'] == "false") and (scores[name]['moon_phase']['value'] != 8)):
            if name not in flagged:
                flagged[name] = ""
            flagged[name] = flagged[name] + "Claimed the moon was not visible but indicated a moon phase from the picture"

        if scores[name]['moon_phase']['value'] in acceptablePhase:
            scores[name]['moon_phase']['score'] = 2
            if scores[name]['moon_phase']['value'] == 8:
                scores[name]['moon_phase']['comment'] = ""
            else:
                scores[name]['moon_phase']['comment'] = f"Accepted {scores[name]['moon_phase']['valueText']} for {fraction_illuminated} illumination"
        else:
            scores[name]['moon_phase']['score'] = 0
            scores[name]['moon_phase']['comment'] = f"Did not accept {scores[name]['moon_phase']['valueText']} for {fraction_illuminated} illumination"
    if verbose:
        print(f" Moon Phase: {scores[name]['moon_phase']['score']} points {scores[name]['moon_phase']['comment']}", file=sys.stderr)

    # 1  Match name with phase (look up)
    scores[name]['moon_phase_name']['value'] = answers.pop(0)
    if not "valueText" in scores[name]['moon_phase']:
        scores[name]['moon_phase_name']['score'] = 0
        scores[name]['moon_phase_name']['comment'] = "Can not look up phase name when previous answer not given"
    else:
        if scores[name]['moon_phase_name']['value'] == scores[name]['moon_phase']['valueText']:
            scores[name]['moon_phase_name']['score'] = 1
            scores[name]['moon_phase_name']['comment'] = ""
        else:
            scores[name]['moon_phase_name']['score'] = 0
            scores[name]['moon_phase_name']['comment'] = f"The number selected in the previous question was {scores[name]['moon_phase']['valueText']}, not {scores[name]['moon_phase_name']['value']}"

    if verbose:
        print(f" Moon Phase Name: {scores[name]['moon_phase_name']['score']} point {scores[name]['moon_phase_name']['comment']}", file=sys.stderr)

    # Location in sky (one point)
    # look up based on moon rise time and moon alt on a given day
    # find max alt and divide sky into quadrants
    # * cutoff separating the lower and upper quadrants determined by $moon_location_cutoff
    # * lower quadrant and azimuth < 180: close to the eastern horizon
    # * upper quadrant and azimuth < 180: in the eastern sky, between the horizon and overhead (the meridian)
    # * max alt, azimutn ~ 180:           overhead (on the meridian) close to overhead (the meridian)
    # * upper quadrant and azimuth > 180: in the western sky, between the horizon and overhead (the meridian)
    # * lower quadrant and azimuth > 180: close to the western horizon
    # * no moon
    # a,b rising
    # d,e setting
    # allow +/- 10 degrees in positions - $moon_location_slop
    scores[name]['moon_location']['value'] = answers.pop(0)
    moon_altitude_percent_of_max = moon_altitude/moon_max_altitude
    if not scores[name]['moon_location']['value'] in skyLocation:
        scores[name]['moon_location']['score'] = 0
        scores[name]['moon_location']['comment'] = f"Answer not a one of the choices - {scores[name]['moon_location']['value']}"
    else:
        acceptableLocation = {}
        # moon_altitude and moon_azimuth will be -1 if it wasn't in the table - not visible!
        if expected_moon_visible == "false":
            acceptableLocation[6] = 1
        else:
            if ((moon_azimuth < 180) and (moon_altitude_percent_of_max < (moon_location_cutoff+moon_location_slop))):
                acceptableLocation[1] = 1 
            if ((moon_azimuth < 180) and (moon_altitude_percent_of_max > (moon_location_cutoff-moon_location_slop))):
                acceptableLocation[2] = 1 
            if (moon_altitude_percent_of_max > (1-moon_location_slop)):
                acceptableLocation[3] = 1 
            if ((moon_azimuth > 180) and (moon_altitude_percent_of_max > (moon_location_cutoff-moon_location_slop))):
                acceptableLocation[4] = 1 
            if ((moon_azimuth > 180) and (moon_altitude_percent_of_max < (moon_location_cutoff+moon_location_slop))):
                acceptableLocation[5] = 1 

        if skyLocation[scores[name]['moon_location']['value']] in acceptableLocation:
            scores[name]['moon_location']['score'] = 1
            if scores[name]['moon_location']['value'] == "no moon":
                scores[name]['moon_location']['comment'] = ""
            else:
                scores[name]['moon_location']['comment'] = ""
                scores[name]['moon_location']['commentTeacher'] = f"Accepted {scores[name]['moon_location']['value']} for percent max alt {moon_altitude_percent_of_max} and azimuth {moon_azimuth}"
        else:
            scores[name]['moon_location']['score'] = 0
            scores[name]['moon_location']['comment'] = "Wanted "
            fencePost = 0
            for my_key in acceptableLocation.keys():
                if fencePost:
                    scores[name]['moon_location']['comment'] = scores[name]['moon_location']['comment'] + "or "
                fencePost = 1
                scores[name]['moon_location']['comment'] = scores[name]['moon_location']['comment'] + f"\"{reverseSkyLocation[my_key]}\" ";
            scores[name]['moon_location']['commentTeacher'] = f"Did not accept {scores[name]['moon_location']['value']} for percent max alt {moon_altitude_percent_of_max} and azimuth {moon_azimuth}"

    if verbose:
        print(f" Moon Location: {scores[name]['moon_location']['score']} point {scores[name]['moon_location']['comment']}", file=sys.stderr)
        print(f"", file=sys.stderr)

    scores[name]['saw_moon']['value'] = 1
    # here we modify the results of the grades on the moon
    # observations in case they say they did not see the moon
    if scores[name]['moon_visible_tf']['value'] == "false":
        scores[name]['saw_moon']['value'] = 0
        if expected_moon_visible == "true":
            # if they didn't answer cuz they thought they were done
            # (<Unanswered> may have been stripped so also checking for empty values:
            no_answers_given_for_moon_quesions = False
            if((scores[name]['moon_phase']['value'] == "<Unanswered>")
               and (scores[name]['moon_phase_name']['value'] == "<Unanswered>")
               and (scores[name]['moon_location']['value'] == "<Unanswered>")):
                no_answers_given_for_moon_quesions = True
            if((scores[name]['moon_phase']['value'] == "")
               and (scores[name]['moon_phase_name']['value'] == "")
               and (scores[name]['moon_location']['value'] == "")):
                no_answers_given_for_moon_quesions = True
            if no_answers_given_for_moon_quesions:
                if observation_number < 3:
                    scores[name]['moonModification']['score'] = 2
                    scores[name]['moonModification']['comment'] = "2 points added to total score so that penalty for missing the moon is 3 points.  Warning: additional point will be deducted after observation 2 for leaving questions unanswered."
                    if verbose:
                        print(f" Moon Fix: Gave gave {scores[name]['moonModification']['score']} points - {scores[name]['moonModification']['comment']}", file=sys.stderr)
                else:
                    scores[name]['moonModification']['score'] = 1
                    scores[name]['moonModification']['comment'] = "1 points added to total score so that penalty for missing the moon and leaving questions unanswered is 4 points."
                    if verbose:
                        print(f" Moon Fix: Gave gave {scores[name]['moonModification']['score']} points - {scores[name]['moonModification']['comment']}", file=sys.stderr)
            # if they correctly answered as if the moon wasn't up
            if((scores[name]['moon_phase']['value'] == 8) and
               (scores[name]['moon_phase_name']['value'] == "no moon") and
               (scores[name]['moon_location']['value'] == "no moon")):
                scores[name]['moonModification']['score'] = 1
                scores[name]['moonModification']['comment'] = "1 point added to total score so that penalty for missing the moon is 3 points."
                if verbose:
                    print(f" Moon Fix: Gave gave {scores[name]['moonModification']['score']} points - {scores[name]['moonModification']['comment']}", file=sys.stderr)
        else:
            moon_entries_blank = 0
            if scores[name]['moon_phase']['value'] == "<Unanswered>" or scores[name]['moon_phase']['value'] == "":
                moon_entries_blank = 1
                scores[name]['moon_phase']['score'] = 2
                scores[name]['moon_phase']['comment'] = "Moon Phase Unaswered, points given, but please fill in next time"
                if verbose:
                    print(f" Moon Phase Fix: gave points even though moon phase was not given", file=sys.stderr)

            if scores[name]['moon_phase_name']['value'] == "<Unanswered>" or scores[name]['moon_phase_name']['value'] == "":
                moon_entries_blank = 1;
                scores[name]['moon_phase_name']['score'] = 1
                scores[name]['moon_phase_name']['comment'] = "Moon Phase Name unaswered, points given, but please fill in next time"
                if verbose:
                    print(f" Moon Phase Name Fix: gave points even though moon phase name was not given", file=sys.stderr)
    
            if scores[name]['moon_location']['value'] == "<Unanswered>" or scores[name]['moon_location']['value'] == "":
                moon_entries_blank = 1;
                scores[name]['moon_location']['score'] = 1
                scores[name]['moon_location']['comment'] = "Moon Location unaswered, points given, but please fill in next time"
                if verbose:
                    print(f" Moon Location Fix: gave points even though moon location was not given", file=sys.stderr)

            if moon_entries_blank:
                if observation_number < 3:
                    scores[name]['moonModification']['score'] = 0
                    scores[name]['moonModification']['comment'] = "After observation 2, one point will be deducted for not answering all moon questions"
                else:
                    scores[name]['moonModification']['score'] = -1
                    scores[name]['moonModification']['comment'] = "One point deducted for not answering all moon questions";
                if verbose:
                    print(f" Moon Fix: Took away {scores[name]['moonModification']['score']} points - {scores[name]['moonModification']['comment']}", file=sys.stderr)

    calculateTotalPoints(scores[name], questions, verbose)



# This subroutine takes a list of answers and grades them
#
# - For each possible question, we fill in value/score/comment (even if the comment is the null string)
# - We optionally fill in a "commentTeacher" parameter which is only output to the teacher
# - we also fill in a "flagged" hash table that is output to the teacher for followon analysis
def gradeSummary(answers, scores, flagged, previous_observation_dates, sunset_times, moon_visible, zeroed_score, myGlobals):
    year = myGlobals["year"]
    # start_date = myGlobals["start_date"]
    # end_date = myGlobals["end_date"]
    # observation_number = myGlobals["observation_number"]
    # questions = myGlobals["questions"]
    # daylight_savings_start = myGlobals["daylight_savings_start"]
    # daylight_savings_check = myGlobals["daylight_savings_check"]
    # moon_altitude_min = myGlobals["moon_altitude_min"]
    # alt_azm_dir = myGlobals["alt_azm_dir"]
    # reverseNamePhase = myGlobals["reverseNamePhase"]
    # skyLocation = myGlobals["skyLocation"]
    verbose = myGlobals["verbose"]
    # sun_degree_slop = myGlobals["sun_degree_slop"]
    # moon_phase_slop = myGlobals["moon_phase_slop"]
    # moon_location_slop = myGlobals["moon_location_slop"]
    # moon_location_cutoff = myGlobals["moon_location_cutoff"]
    # reverseSkyLocation = myGlobals["reverseSkyLocation"]
    spring_semester = myGlobals["spring_semester"]
    summary_degrees_slop = myGlobals["summary_degrees_slop"]
    
    name = answers.pop(0)
    scores[name]["lastName"] = answers.pop(0);
    scores[name]["firstName"] = answers.pop(0);
    scores[name]["blackboardEntries"] = answers.copy();
    if verbose:
        print(f"{scores[name]['lastName']}, {scores[name]['firstName']} ({name})", file=sys.stderr)
        print(f" Blackboard Entries: {scores[name]['blackboardEntries']}", file=sys.stderr)
    my_last_observation_date = datetime.datetime.strptime(myGlobals["summary_all"][name]["LastObservationDate"], '%Y-%m-%d %H:%M:%S')
    official_last_month = my_last_observation_date.month
    official_last_day = my_last_observation_date.day

    # Question 1
    question=1
    scores[name][question]['value'] = answers.pop(0);
    print(f"Question {question}:\n\tResponse:{scores[name][question]['value']}", file=sys.stderr)
    scores[name][question]['score'] = 0
    scores[name][question]['comment'] = f"Question {question} incorrect"
    if spring_semester:
        if scores[name][question]["value"] == myGlobals["summary_questions_01"][1]:
            scores[name][question]["score"] = 2;
            scores[name][question]['comment'] = f"Question {question} correct"
    else:
        if scores[name][question]["value"] == myGlobals["summary_questions_01"][2]:
            scores[name][question]["score"] = 2;
            scores[name][question]['comment'] = f"Question {question} correct"
    print(f"{scores[name][question]['comment']}", file=sys.stderr)
    
    # Question 2
    question += 1
    scores[name][question]["value"] = answers.pop(0);
    scores[name][question]["value"] = scores[name][question]["value"].lower()
    print(f"Question {question}:\n\tResponse:{scores[name][question]['value']}", file=sys.stderr)
    scores[name][question]["score"] = 0
    scores[name][question]['comment'] = f"Question {question} incorrect"
    if scores[name][question]["value"] == "true":
        scores[name][question]["score"] = 2
        scores[name][question]['comment'] = f"Question {question} correct"
    print(f"{scores[name][question]['comment']}", file=sys.stderr)

    # Question 3
    question += 1;
    scores[name][question]["value"] = answers.pop(0);
    count_q3_answers_array  = scores[name][question]["value"].split(",")
    count_q3_answers = len(count_q3_answers_array)
    print(f"Question {question}:\n\tResponse:{scores[name][question]['value']}", file=sys.stderr)
    scores[name][question]["score"] = 0
    scores[name][question]["comment"] = f"There were 2 correct answers, you gave {count_q3_answers} answers.  Your correct answers are listed below:\n";
    # Note: for spring/fall changes Spring(2 and 6) Fall (1,7)
    if spring_semester:
        if myGlobals['summary_questions_03'][2] in scores[name][3]["value"]:
            scores[name][3]["score"] += 1
            print(f"Question 3 - got {myGlobals['summary_questions_03'][2]} correct\n", file=sys.stderr)
            scores[name][3]["comment"] = scores[name][3]["comment"] + f"\t\t  - got {myGlobals['summary_questions_03'][2]} correct\n"
        if myGlobals['summary_questions_03'][6] in scores[name][3]["value"]:
            scores[name][3]["score"] += 1
            print(f"Question 3 - got {myGlobals['summary_questions_03'][6]} correct\n", file=sys.stderr)
            scores[name][3]["comment"] = scores[name][3]["comment"] + f"\t\t  - got {myGlobals['summary_questions_03'][6]} correct\n"
    else:
        if myGlobals['summary_questions_03'][1] in scores[name][3]["value"]: 
            scores[name][3]["score"] += 1
            print(f"Question 3 - got {myGlobals['summary_questions_03'][1]} correct\n", file=sys.stderr)
            scores[name][3]["comment"] = scores[name][3]["comment"] + f"\t\t  - got {myGlobals['summary_questions_03'][1]} correct\n"
        if myGlobals['summary_questions_03'][7] in scores[name][3]["value"]:
            scores[name][3]["score"] += 1
            print(f"Question 3 - got {myGlobals['summary_questions_03'][7]} correct\n", file=sys.stderr)
            scores[name][3]["comment"] = scores[name][3]["comment"] + f"\t\t  - got {myGlobals['summary_questions_03'][7]} correct\n"
    if count_q3_answers > 2:
        points_to_remove = count_q3_answers-2
        scores[name][3]["score"] -= points_to_remove
        if scores[name][3]["score"] < 0:
            scores[name][3]["score"] = 0
        scores[name][3]["comment"] = scores[name][3]["comment"] + f"\t\t  WARNING: Gave more than 2 answers - docked {points_to_remove} point(s)."
    print(f"{scores[name][question]['comment']}", file=sys.stderr)

    # Question 4
    question += 1
    scores[name][question]["value"] = answers.pop(0);
    print(f"Question {question}:\n\tResponse:{scores[name][question]['value']}", file=sys.stderr)
    try:
        scores[name][question]["value"] = int(scores[name][question]["value"])
    except:
        scores[name][question]["value"] = ""
    scores[name][question]["score"] = 0
    scores[name][question]['comment'] = f"Incorrect last observation month - does {scores[name][question]['value']} equal {official_last_month}?"
    if scores[name][question]["value"] == official_last_month:
        scores[name][question]["score"] = 1
        scores[name][question]['comment'] = f"Question {question} correct"
    print(f"{scores[name][question]['comment']}", file=sys.stderr)

    # Question 5
    question += 1
    scores[name][question]["value"] = answers.pop(0)
    print(f"Question {question}:\n\tResponse:{scores[name][question]['value']}", file=sys.stderr)
    scores[name][question]["score"] = 0
    try:
        scores[name][question]["value"] = int(scores[name][question]["value"])
    except:
        scores[name][question]["value"] = ""
    scores[name][question]["comment"] = f"Incorrect last observation day - does {scores[name][question]['value']} equal {official_last_day}?"
    if scores[name][question]["value"] == official_last_day:
        scores[name][question]["score"] = 1
        scores[name][question]['comment'] = f"Question {question} correct"
    if myGlobals["summary_all"][name][ myGlobals['summary_all'][name]['LastObservationDate'] ]['ObservationWasZeroed'] != "0":
        scores[name][question]["comment"] = scores[name][question]["comment"] + "WARNING: Last observation values were zeroed due to incorrect observation time"
    print(f"{scores[name][question]['comment']}", file=sys.stderr)

    # Question 6
    question += 1
    scores[name][question]["value"] = answers.pop(0)
    scores[name][question]["value"] = scores[name][question]["value"].lower()
    print(f"Question {question}:\n\tResponse:{scores[name][question]['value']}", file=sys.stderr)
    scores[name][question]['score'] = 0 
    scores[name][question]['comment'] = f"Question {question} incorrect"
    # CRAIG - fix this for spring/fall changes Spring(False) Fall (True)
    if spring_semester and (scores[name][question]["value"] == "false"):
        scores[name][question]["score"] = 2
        scores[name][question]['comment'] = f"Question {question} correct"
    if (not spring_semester) and (scores[name][question]["value"] == "true"):
        scores[name][question]["score"] = 2
        scores[name][question]['comment'] = f"Question {question} correct"
    print(f"{scores[name][question]['comment']}", file=sys.stderr)

    # Question 7
    # DOME - is this really a question that nobody can answer sometimes?
    question += 1
    official_last_degrees = myGlobals['summary_all'][name][ myGlobals['summary_all'][name]['LastObservationDate'] ]['DegreesNorth']
    scores[name][question]["value"] = answers.pop(0)
    print(f"Question {question}:\n\tResponse:{scores[name][question]['value']}", file=sys.stderr)
    if not is_float(scores[name][question]["value"]):
        scores[name][question]["value"] = re.sub(r'.*?(\d+).*', r'\1', scores[name][question]["value"])
        print(f"Question {question}:\tCleaned to {scores[name][question]['value']}", file=sys.stderr)
    try:
        scores[name][question]["value"] = float(scores[name][question]["value"])
    except:
        scores[name][question]["value"] = ""
    scores[name][question]['score'] = 0 
    if official_last_degrees == -1:
        scores[name][question]["comment"] = "Saved degrees from last observation = None"
    elif not scores[name][question]["value"]:
        scores[name][question]["comment"] = "Number not entered correctly"
    else:
        scores[name][question]["comment"] = f"Saved degrees from last observation = {official_last_degrees}"
        if abs(scores[name][question]["value"]-official_last_degrees) < summary_degrees_slop:
            scores[name][question]["score"] = 2
            print(f"Question {question} correct", file=sys.stderr)
    print(f"{scores[name][question]['comment']}", file=sys.stderr)

    # Question 8
    question += 1
    scores[name][question]["value"] = answers.pop(0)
    scores[name][question]["value"] = scores[name][question]["value"].lower()
    print(f"Question {question}:\n\tResponse:{scores[name][question]['value']}", file=sys.stderr)
    scores[name][question]["score"] = 0
    scores[name][question]['comment'] = f"Question {question} incorrect"
    # CRAIG - fix this for spring/fall changes Spring(True) Fall (False)
    if spring_semester and (scores[name][question]["value"] == "true"):
        scores[name][question]["score"] = 2
        scores[name][question]['comment'] = f"Question {question} correct"
    if (not spring_semester) and (scores[name][question]["value"] == "false"):
        scores[name][question]["score"] = 2
        scores[name][question]['comment'] = f"Question {question} correct"
    print(f"{scores[name][question]['comment']}", file=sys.stderr)

    # Question 9
    question += 1
    scores[name][question]["value"] = answers.pop(0)
    print(f"Question {question}:\n\tResponse:{scores[name][question]['value']}", file=sys.stderr)
    if not scores[name][question]["value"].isdecimal():
        scores[name][question]["value"] = re.sub(r'.*?(\d+).*', r'\1', scores[name][question]["value"])
        print(f"Question {question}:\tCleaned to {scores[name][question]['value']}", file=sys.stderr)
    try:
        scores[name][question]["value"] = float(scores[name][question]["value"])
    except:
        scores[name][question]["value"] = ""
    scores[name][question]["score"] = 0
    if official_last_degrees == -1:
        scores[name][question]["comment"] = "Saved degrees from last observation = None"
    elif not scores[name][question]["value"]:
        scores[name][question]["comment"] = "Number not entered correctly"
    else:
        scores[name][question]["comment"] = f"Saved degrees from last observation = {official_last_degrees}"
        if abs(scores[name][question]["value"]-official_last_degrees) < summary_degrees_slop:
            scores[name][question]["score"] = 2
            print(f"Question {question} correct", file=sys.stderr)
    print(f"{scores[name][question]['comment']}", file=sys.stderr)

    # Question 10
    question += 1
    date_two_weeks_ahead = my_last_observation_date + datetime.timedelta(days=14)
    acceptablePhase = calculate_acceptable_phases(date_two_weeks_ahead, myGlobals)
    scores[name][question]["value"] = answers.pop(0)
    print(f"Question {question}:\n\tResponse:{scores[name][question]['value']}", file=sys.stderr)
    scores[name][question]["score"] = 0
    scores[name][question]["comment"] = ""
    question_10_answer = ""
    if not scores[name][question]["value"] in myGlobals['summaryNamePhase']:
        print(f"WARNING - Question 10 should have matched something", file=sys.stderr)
        scores[name][question]["comment"] = scores[name][question]["comment"] + "WARNING - Question 10 should have matched something\n"
    else:
        question_10_answer = myGlobals['summaryNamePhase'][scores[name][question]["value"]]
    y = date_two_weeks_ahead.year
    m = date_two_weeks_ahead.month
    d = date_two_weeks_ahead.day
    print(f"Date: {m}/{d}/{y}", file=sys.stderr)
    scores[name][question]["comment"] = scores[name][question]["comment"] + f"Date: {m}/{d}/{y}"
    scores[name][question]["comment"] = scores[name][question]["comment"] + f"\n\t\t  {acceptablePhase['comment']}"

    for my_key in range(1,9):
        if my_key in acceptablePhase:
            print(f"Accept phase: {myGlobals['summaryReverseNamePhase'][my_key]}", file=sys.stderr)
            scores[name][question]["comment"] = scores[name][question]["comment"] + f"\n\t\t  Acceptable: {myGlobals['summaryReverseNamePhase'][my_key]}" 
    if question_10_answer in acceptablePhase:
            scores[name][question]["score"] = 2
            print(f"Question 10 correct", file=sys.stderr)

    # Question 11
    question += 1
    scores[name][question]["value"] = answers.pop(0)
    print(f"Question {question}:\n\tResponse:{scores[name][question]['value']}", file=sys.stderr)
    scores[name][question]["score"] = 0
    question_11_answer = ""
    question_11_correct_answers = {}
    question_11_correct_answers[0] = 0
    question_11_correct_answers[1] = 1
    question_11_correct_answers[2] = 2
    question_11_correct_answers[3] = 3
    question_11_correct_answers[4] = 4
    question_11_correct_answers[5] = 3
    question_11_correct_answers[6] = 2
    question_11_correct_answers[7] = 1
    question_11_correct_answers[8] = 0
    for ii in range(len(myGlobals['summary_questions_11'])):
        if scores[name][question]["value"] == myGlobals['summary_questions_11'][ii]:
            question_11_answer = ii 
    if question_11_answer ==  "":
        scores[name][question]["comment"] = f"WARNING - Question 11 should have matched something\n"
    else:
        if question_10_answer in question_11_correct_answers:
            scores[name][question]["comment"] = f"Question 11: Based on Question 10 answer should be {myGlobals['summary_questions_11'][question_11_correct_answers[question_10_answer]]}"
            if question_11_answer == question_11_correct_answers[question_10_answer]:
                scores[name][question]["score"] = 2
                scores[name][question]["comment"] = scores[name][question]["comment"] + f"\n\t\t  Question 11 correct"
        else:
            scores[name][question]["comment"] = f"Question 11: Bad answer for question 10 so no match"
    print(scores[name][question]["comment"], file=sys.stderr)

    # Question 12
    question += 1
    date_one_month_ahead = my_last_observation_date + datetime.timedelta(days=28)
    acceptablePhase = calculate_acceptable_phases(date_one_month_ahead, myGlobals)
    scores[name][question]["value"] = answers.pop(0)
    print(f"Question {question}:\n\tResponse:{scores[name][question]['value']}", file=sys.stderr)
    scores[name][question]["score"] = 0
    scores[name][question]["comment"] = ""
    question_12_answer = ""
    if not scores[name][question]["value"] in myGlobals["summaryNamePhase"]:
        scores[name][question]["comment"] = "WARNING - Question 12 should have matched something\n"
    else:
        question_12_answer = myGlobals["summaryNamePhase"][scores[name][question]["value"]]
    y = date_one_month_ahead.year
    m = date_one_month_ahead.month
    d = date_one_month_ahead.day
    scores[name][question]["comment"] = scores[name][question]["comment"] + f"Date: {m}/{d}/{y}"
    scores[name][question]["comment"] = scores[name][question]["comment"] + f"\n\t\t  {acceptablePhase['comment']}"
    for my_key in range(1,9):
        if my_key in acceptablePhase:
            scores[name][question]["comment"] = scores[name][question]["comment"] + f"\n\t\t  Acceptable: {myGlobals['summaryReverseNamePhase'][my_key]}"  
    if question_12_answer in acceptablePhase:
        scores[name][question]["score"] = 2
        scores[name][question]["comment"] = scores[name][question]["comment"] + f"\n\t\t  Question 12 correct"
    print(scores[name][question]["comment"], file=sys.stderr)

    # Question 13
    question += 1
    scores[name][question]["value"] = answers.pop(0)
    print(f"Question {question}:\n\tResponse:{scores[name][question]['value']}", file=sys.stderr)
    scores[name][question]["score"] = 0
    question_13_answer = ""
    question_13_correct_answers = {}
    question_13_correct_answers[0] = 0
    question_13_correct_answers[1] = 1
    question_13_correct_answers[2] = 2
    question_13_correct_answers[3] = 3
    question_13_correct_answers[4] = 4
    question_13_correct_answers[5] = 3
    question_13_correct_answers[6] = 2
    question_13_correct_answers[7] = 1
    question_13_correct_answers[8] = 0
    for ii in range(len(myGlobals['summary_questions_13'])):
        if scores[name][question]["value"] == myGlobals['summary_questions_13'][ii]:
            question_13_answer = ii 
    if question_13_answer == "":
        scores[name][question]["comment"] = f"WARNING - Question 13 should have matched something"
    else:
        if not question_12_answer in question_13_correct_answers:
            scores[name][question]["comment"] = f"Question 13: Couldn't parse Question 12 so unable to answer Question 13"
        else:
            scores[name][question]["comment"] = f"Question 13: Based on Question 12 answer should be {myGlobals['summary_questions_13'][ question_13_correct_answers[question_12_answer] ]}"
            if question_13_answer == question_13_correct_answers[question_12_answer]:
                scores[name][question]["score"] = 2
                scores[name][question]["comment"] = scores[name][question]["comment"] + f"\n\t\t  Question 13 correct"
    print(scores[name][question]["comment"], file=sys.stderr)

    # Question 14 and Question 15
    question += 1
    scores[name][question]["value"] = answers.pop(0)
    print(f"Question {question}:\n\tResponse:{scores[name][question]['value']}", file=sys.stderr)
    if not scores[name][question]["value"].isdecimal():
        scores[name][question]["value"] = re.sub(r'.*?(\d+).*', r'\1', scores[name][question]["value"]).zfill(2)
        print(f"Question {question}:\tCleaned to {scores[name][question]['value']}", file=sys.stderr)
    scores[name][question]["value"] = int(scores[name][question]["value"])
    scores[name][question]['score'] = 0
    scores[name][question]["comment"] = ""    
    question += 1
    scores[name][question]["value"] = answers.pop(0)
    print(f"Question {question}:\n\tResponse:{scores[name][question]['value']}", file=sys.stderr)
    if not scores[name][question]["value"].isdecimal():
        scores[name][question]["value"] = re.sub(r'.*?(\d+).*', r'\1', scores[name][question]["value"])
        print(f"Question {question}:\tCleaned to {scores[name][question]['value']}", file=sys.stderr)
    try:
        scores[name][question]["value"] = int(scores[name][question]["value"])
    except:
        scores[name][question]["value"] = ""
    scores[name][question]['score'] = 0
    scores[name][question]["comment"] = ""    

    # Check that they found one date where they saw the moon
    last_no_moon_date = ""
    no_moon_date = ""

    # some of these keys are not actually observation dates - we check that implicitly below
    for obs_date in myGlobals['summary_all'][name].keys():
        if(obs_date == "LastObservationDate"):
            continue 
        if ('SawMoon' in myGlobals['summary_all'][name][obs_date]) and (myGlobals['summary_all'][name][obs_date]['SawMoon'] == 0):
            # mark that there was at least one date where
            last_no_moon_date = datetime.datetime.strptime(obs_date, '%Y-%m-%d %H:%M:%S')
            month = last_no_moon_date.month
            day = last_no_moon_date.day
            if (month == scores[name][14]["value"]) and (day == scores[name][15]["value"]):
                no_moon_date = last_no_moon_date
    if last_no_moon_date and (not no_moon_date):
        scores[name][14]['score'] = 0
        scores[name][15]['score'] = 0
        last_no_moon_date_month_day = f"{month}/{day}"
        print(f"No Points for Questions 14 and 15", file=sys.stderr)
        print(f"- They said they did not see the moon on {last_no_moon_date_month_day}, but they input {scores[name][14]['value']}/{scores[name][15]['value']}", file=sys.stderr)
        scores[name][14]["comment"] = "No Points for Questions 14 and 15\n"
        scores[name][14]["comment"] = scores[name][14]["comment"] + f"- You did not see the moon on {last_no_moon_date_month_day}, but you input {scores[name][14]['value']}/{scores[name][15]['value']}\n"
        scores[name][15]["comment"] = scores[name][14]["comment"]
        date = f'{scores[name][14]["value"]}/{scores[name][15]["value"]}/{year}'
        try:
            no_moon_date = datetime.datetime.strptime(date, '%m/%d/%Y')
        except:
            no_moon_date = ""
    else:
        scores[name][14]["score"] = 1
        scores[name][15]["score"] = 1
        print(f"Correct Questions 14 and 15", file=sys.stderr)
        if myGlobals['summary_all'][name][no_moon_date]['MoonShouldHaveBeenVisible']:
            scores[name][15]['comment'] = "WARNING: Moon should have been visible on this date\n"
    print(scores[name][14]["comment"], file=sys.stderr)
    print(scores[name][15]["comment"], file=sys.stderr)

    # Question 16
    question += 1
    scores[name][question]["value"] = answers.pop(0)
    print(f"Question {question}:\n\tResponse:{scores[name][question]['value']}", file=sys.stderr)
    scores[name][question]["score"] = 0
    scores[name][question]["comment"] = ""
    question_16_answer = ""
    if scores[name][question]["value"] in myGlobals["summaryNamePhase"]:
        question_16_answer = myGlobals["summaryNamePhase"][ scores[name][question]["value"] ]
    if not no_moon_date:
        scores[name][question]["comment"] = f"Ignoring Question 16 - could not parse date given in previous question"
    else:
        acceptablePhase = calculate_acceptable_phases(no_moon_date, myGlobals)
        if question_16_answer == "":
            scores[name][question]["comment"] = f"WARNING - Question 16 should have matched something\n"
        y = no_moon_date.year
        m = no_moon_date.month
        d = no_moon_date.day
        scores[name][question]["comment"] = scores[name][question]["comment"] + f"Date: {m}/{d}/{y}"
        scores[name][question]["comment"] = scores[name][question]["comment"] + f"\n\t\t  {acceptablePhase['comment']}"
        for my_key in range(1,9):
            if my_key in acceptablePhase:
                scores[name][question]["comment"] = scores[name][question]["comment"] + f"\n\t\t  Acceptable: {myGlobals['summaryReverseNamePhase'][my_key]}" 
        if question_16_answer in acceptablePhase:
            scores[name][question]["score"] = 2
            scores[name][question]["comment"] = scores[name][question]["comment"] + "\n\t\t  Question 16 correct"
    print(scores[name][question]["comment"], file=sys.stderr)

    # Question 17
    question += 1
    scores[name][question]["value"] = answers.pop(0)
    print(f"Question {question}:\n\tResponse:{scores[name][question]['value']}", file=sys.stderr)
    scores[name][question]["score"] = 0
    scores[name][question]["comment"] = ""
    question_17_answer = ""
    question_17_correct_answers = {}
    question_17_correct_answers[0] = 0
    question_17_correct_answers[1] = 1
    question_17_correct_answers[2] = 2
    question_17_correct_answers[3] = 3
    question_17_correct_answers[4] = 4
    question_17_correct_answers[5] = 3
    question_17_correct_answers[6] = 2
    question_17_correct_answers[7] = 1
    question_17_correct_answers[8] = 0
    for ii in range(len(myGlobals['summary_questions_17'])):
        if scores[name][question]["value"] == myGlobals['summary_questions_17'][ii]:
           question_17_answer = ii 
    if question_17_answer == "":
        scores[name][question]["comment"] = "WARNING - Question 17 should have matched something"
    else:
        if question_16_answer not in question_17_correct_answers:
            scores[name][question]["comment"] = f"Question 17: Bailing because Question 16 answer was not understood."
        else:
            scores[name][question]["comment"] = f"Question 17: Based on Question 16 answer should be {myGlobals['summary_questions_17'][ question_17_correct_answers[question_16_answer] ]}"
            if question_17_answer == question_17_correct_answers[question_16_answer]:
                scores[name][question]["score"] = 2
                scores[name][question]["comment"] = f"\n\t\t  Question 17 correct"
    print(scores[name][question]["comment"], file=sys.stderr)

    # Question 18
    question += 1
    scores[name][question]["value"] = answers.pop(0)
    count_q18_answers_array  = scores[name][question]["value"].split(",")
    count_q18_answers = len(count_q18_answers_array)
    print(f"Question {question}:\n\tResponse:{scores[name][question]['value']}", file=sys.stderr)
    scores[name][question]["score"] = 0
    scores[name][question]["comment"] = f"There were 4 correct answers, you gave {count_q18_answers} answers.  Your correct answers are listed below:"
    for answer in count_q18_answers_array:
        answer = answer.strip()
        if answer == myGlobals['summary_questions_18'][0]:
            scores[name][question]['score'] += .5
            scores[name][question]["comment"] = scores[name][question]["comment"] + f"\n\t\t  - got {myGlobals['summary_questions_18'][0]} correct"
        elif answer == myGlobals['summary_questions_18'][5]:
            scores[name][question]['score'] += .5
            scores[name][question]["comment"] = scores[name][question]["comment"] + f"\n\t\t  - got {myGlobals['summary_questions_18'][5]} correct"
        elif answer == myGlobals['summary_questions_18'][6]:
            scores[name][question]['score'] += .5
            scores[name][question]["comment"] = scores[name][question]["comment"] + f"\n\t\t  - got {myGlobals['summary_questions_18'][6]} correct"
        elif answer == myGlobals['summary_questions_18'][7]:
            scores[name][question]['score'] += .5
            scores[name][question]["comment"] = scores[name][question]["comment"] + f"\n\t\t  - got {myGlobals['summary_questions_18'][7]} correct"
    if count_q18_answers > 4:
        points_to_remove = 0.5 * (count_q18_answers - 4)
        scores[name][question]['score'] -= points_to_remove
        if scores[name][question]['score'] < 0:
            scores[name][question]['score'] = 0
        scores[name][question]["comment"] = scores[name][question]["comment"] + f"\n\t\t  WARNING: Gave more than 4 answers - docked {points_to_remove} point(s)."
    print(scores[name][question]["comment"], file=sys.stderr)

    # Question 19
    question += 1
    scores[name][question]["value"] = answers.pop(0)
    scores[name][question]["score"] = 0
    print(f"Question {question}:\n\tResponse:{scores[name][question]['value']}", file=sys.stderr)
    scores[name]["totalPoints"] = 0
    for ii in range(1, 19):
        scores[name]["totalPoints"] += scores[name][ii]["score"]
    print(f"{name} got {scores[name]['totalPoints']} total points", file=sys.stderr)
    print(f"Summay Done --------------------", file=sys.stderr)
